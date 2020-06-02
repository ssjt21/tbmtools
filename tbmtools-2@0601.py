# -*- coding:utf-8 -*-
'''
@author = 随时静听 - xuqiu
@date = 2020-05-23 17:48
@product_name = PyCharm
@project_name = KeySearch
@filename = tmbDataInput.py
'''

import os
import re
from argparse import ArgumentParser
import openpyxl
import json
import glob
import string
from datetime import datetime

#取数据 4 匹配版本
# 5 获取来源 空的填写
# 8 ，9
# 行业小类 7
'''
{
"apache":{"2.2":[],"2.3":[]},

}
'''

def getVersion(sheet):
    if not sheet:
        return

    version=sheet.cell(row=2,column=5).value
    row=3
    while not version.strip():
        row=+1
        version=sheet.cell(row=row,column=5).value
    # print version
    if not version.strip():
        return
    version=version.strip()
    pattern=re.search("^([a-zA-Z]+?)(\d+\.\d+)",version)
    if pattern:
        class_type = pattern.group(1)
        ver = pattern.group(2)
        return class_type, ver
    #单个版本号的情况
    pattern=re.search("^([a-zA-Z]+?)(\d+)",version)
    if pattern:
        class_type = pattern.group(1)
        ver = pattern.group(2)
        return class_type, ver

    # 无版本号的情况
    pattern = re.search("^([a-zA-Z]+)", version)
    # print pattern
    if pattern:
        class_type = pattern.group(1)
        print class_type
        return class_type, "noversion"
    return None






def readXlsMap(filename="Apache2.2.xlsx",t=True):
    # if not os.path.exists(filename):
    #     return []
    if not os.path.exists(filename):
        return False
    try:
        wb= openpyxl.load_workbook(filename)
    # sheets = wb.sheetnames
    except:
        print ("[-] XLSX load Failed" + filename)
        return False

    sheet=wb.get_sheet_by_name(wb.sheetnames[0])

    # 获取版本和类型
    cls_ver=getVersion(sheet)

    kv={}
    rowlst=[]
    for i,row in enumerate(sheet.rows):
        if i==0:
            continue
        # print row[7].value
        if row[7].value and kv.has_key(row[7].value):
            continue

        if row[7].value :
            if t:
                if row[0].value:
                    rowlst.append([c.value if c else "" for c in row ])
            else:
                kv.update({row[7].value:[row[5].value, row[8].value,row[9].value]})


    if t:
        return {cls_ver[0]:{cls_ver[1]:rowlst}}

    else:

        return  {cls_ver[0]:{cls_ver[1]:kv}}


## 更新json 映射关系
def writeMap(kv,filename="./data/data.json"):
    try:
        dirname=os.path.dirname(os.path.abspath(filename))
        if not os.path.exists(dirname):
            print "[Error] dirctory is not exists!" + dirname
            return

        dataDict = {}
        if os.path.exists(filename):
            with open(filename,"r") as f:
                if os.path.getsize(filename)!=0:
                    # print 22
                    dataDict=json.load(f,encoding='utf-8')#,strict=False)#, strict=False)

        with open(filename,"w") as f:
            k = list(kv.keys())[0]
            olddict = dataDict.get(k, None)
            if olddict:
                olddict.update(kv[k])
                dataDict.update({k: olddict})
            else:

                dataDict.update(kv)
            json.dump(dataDict, f, indent=2)
        return True,"[Info] JSON map file update success!"+filename
    except Exception ,e:
        return False,"[Error] JSON MAP file load or dump failed! "+e.message

##获取文件列表
def getAllfile(path="templates",ext=".xlsx"):
    if os.path.exists(path) and os.path.isdir(path):
        fileLst=glob.glob1(path,"*"+ext)
        fileLst= map( lambda _:os.path.join(path,_),fileLst)
        return fileLst
    else:
        print("[Error] Get xlsx file failed! Path is not exists!"+path)
        return []

#根据模板路径中的xlsx进行更新json 数据中的映射关系
def upDatejsonmap(tplpath="templates",jsonfile="./data/data.json",on=True):
    if on:
        try:
            fileLst=getAllfile(tplpath)
            for xlsxfile in fileLst:
                dataDict=readXlsMap(xlsxfile,False)
                ret=writeMap(dataDict,jsonfile)
                if ret:
                    print ret[-1]+" for "+xlsxfile
                else:
                    print ret[-1] +" for " +xlsxfile
            return True
        except:
            return  False





# 写入文件xlsx - 弃用，因为小项不唯一

def writeXlsx_old(savefilename,dataDict,jsonMap):


    keys=dataDict.keys()
    if not keys:
        return False
    key=keys[0]
    versions=dataDict[key].keys()
    if not versions:
        return False
    version = versions[0]
    if jsonMap.has_key(key) and jsonMap[key].has_key(version):
        mapValues=jsonMap[key][version]
        values=dataDict[key][version]
        for k,v in values.items():

            if mapValues.has_key(k):
                values[k][5]=u"工信部："+mapValues[k][0] #填充检查标准来源

                values[k][7]=u"检查组补充检查项（可选）"+mapValues[k][1] #填充要求内容
                values[k][8]=mapValues[k][2] #填充操作指南
        # filename = os.path.join(savepath, key+"_"+version+datetime.strftime(datetime.now(), " '%Y%m%d_%H%M%S'")+".xlsx")
        # filename=
        wb = openpyxl.Workbook()
        ws = wb.active
        r=2
        for i,row in values.items():
            for ci,c in enumerate(row):
                print ci
                if ci==0:
                    ws.cell(row=r,column=ci+1,value=r-1)
                else:
                    ws.cell(row=r,column=ci+1,value=c)
            r=r+1
        wb.save(savefilename)
        return True,savefilename
    else:
        return False

def writeXlsx(savefilename,dataDict,jsonMap,titles="./data/title.conf"):


    keys=dataDict.keys()
    if not keys:
        return False
    key=keys[0]
    versions=dataDict[key].keys()
    if not versions:
        return False
    version = versions[0]
    if jsonMap.has_key(key) and jsonMap[key].has_key(version):
        mapValues=jsonMap[key][version]
        values=dataDict[key][version]
        for i, row in enumerate(values):
            k=row[7]
            # print k
            if mapValues.has_key(k):
                if mapValues[k][0]:
                    values[i][5] = u"工信部:" + mapValues[k][0]  # 填充检查标准来源
                else:
                    continue
                # else:
                #     values[i][5] = u"检查组补充检查项（可选）"
                # if not mapValues[k][1]:
                #     values[i][8] = u"检查组补充检查项（可选）"  # 填充要求内容
                # else:
                values[i][8] = mapValues[k][1]  # 填充要求内容
                if mapValues[k][2]:
                    values[i][9] = mapValues[k][2]  # 填充操作指南
                # else:
                #     values[i][9] = u"检查组补充检查项（可选）"
        # filename = os.path.join(savepath, key+"_"+version+datetime.strftime(datetime.now(), " '%Y%m%d_%H%M%S'")+".xlsx")
        # filename=
        wb = openpyxl.Workbook()
        ws = wb.active
        ws=setTitle(ws,titles)
        r=2
        border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin",color ="000000"), \
                right = openpyxl.styles.Side(border_style="thin", color ="000000"),\
                top = openpyxl.styles.Side(border_style="thin", color ="000000"),\
                bottom = openpyxl.styles.Side(border_style="thin", color ="000000"))

        for i,row in enumerate(values):
            for ci,c in enumerate(row):
                if row[5].strip()=="":
                    r=r-1
                    break
                if row[12]!=u"不合规" :
                    r=r-1
                    break

                if ci==0:
                    ws.cell(row=r,column=ci+1,value=r-1)
                    ws.cell(row=r, column=ci + 1, value=r - 1).border=border
                else:
                    ws.cell(row=r,column=ci+1,value=c)
                    ws.cell(row=r,column=ci+1,value=c).border=border
            r=r+1

        wb.save(savefilename)
        return True,savefilename
    else:
        return False

# 读取配置文件写入头并设置格式
def setTitle(ws,conf="./data/title.conf"):
    uppercase=string.uppercase
    style={"bold":True,"italic":True}
    if os.path.exists(conf):
        with open(conf,"r") as f:
            titles=f.read()
            titles=titles.split("\n")
            titles=filter( lambda x:x.strip(),titles)
            titles=map(lambda x:x.split("::"),titles)
            for i,line in enumerate(titles):
                # print line
                ws.cell(row=1,column=i+1,value=line[1])
                #字体样式
                fontstryle=openpyxl.styles.Font(name=u"宋体",size=int(line[2]),bold=style.get(line[3].lower(),False), italic=style.get(line[4].lower(),False),color=line[4]\
                                         )
                ws[uppercase[i]+"1"].font = fontstryle
                ws[uppercase[i] + "1"]=line[1]
                ws[uppercase[i] + "1"].fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=line[5])
                #对齐方式
                ws[uppercase[i]+"1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                #设置行高和列宽
                ws.row_dimensions[1].height=int(line[6])
                ws.column_dimensions[uppercase[i]].width=int(line[7])
    return  ws


def main():
    parser = ArgumentParser(
        description=
        '''TBM Data fill tool! Help:xxxx
        Usage: python tbmtools.py -h
        '''
    )

    parser.add_argument(
        '-m',
        '--make',
        action="store_true",
        help=u'根据模板文件重新生成映射关系文件 data.json')
    parser.add_argument(
        '-t',
        '-templates',
        dest='templates',
        nargs='?',
        default="./templates/",
        help=u'指定生成data.json映射关系的xlsx模板路径,data.json中的映射关系由该目录中的文件生成,默认路径为./templates')
    parser.add_argument(
        '-o',
        '--output',
        dest='output',
        nargs='?',
        default='./reports/',
        help=u'指定导出文件的路径，默认值为 ./reports/'
    )
    parser.add_argument(
        '-i',
        '--input',
        dest='input',
        nargs='?',
        default="",
        help=u'指定需要处理的多个文件路径'
    )
    args = parser.parse_args()
    jsonfile = "./data/data.json"
    if args.make:
        ret=upDatejsonmap(tplpath=args.templates,jsonfile=jsonfile,on=args.make)
        if ret:
            print "[Info] data.json map make successfull!"
        else:
            print "[Error] data.json map make failed!"
            exit()
    if not args.input:
        exit()

    filelst=getAllfile(args.input)
    print "[Info] Find xlsx file num:"+str(len(filelst))

    if os.path.exists(jsonfile):
        try:
            with open(jsonfile, "r") as f:
                jsonmap=json.load(f)
        except Exception,e:
            print "[Error] ./data/data.json 文件加载失败，请检查文件格式是否正确"
            exit()
    else:
        print u"[Error] ./data/data.json 文件不存在，请使用-m 参数生成data.json映射关系文件."
        exit()
    failedfiles=[]
    for filename in filelst:
        dataDict = readXlsMap(filename)
        if dataDict :
            savefilename=os.path.basename(filename)
            savefilename=os.path.splitext(savefilename)[0]
            class_type=dataDict.keys()[0]
            version=dataDict[class_type].keys()[0]
            savefilename=class_type.encode("gbk")+"_".encode("gbk")+version.encode("gbk")+"_".encode("gbk")+savefilename+".xlsx".encode("gbk")
            savefilename=os.path.join(args.output,savefilename)
            ret =writeXlsx(savefilename,dataDict,jsonmap)
            if ret:
                print "[Info] "+filename+" -->"+savefilename+"  success."
            else:
                print "[Failed] " + filename + " -->" + savefilename + "  Failed."
                failedfiles.append(filename)
        else:
            print("[Failed] " + filename + "is bad!")
            failedfiles.append(filename)

    with open("failed.txt","a+")  as f:
        for filename in failedfiles:
            f.write(filename+"\n")
    total=len(filelst)
    failednum=len(failedfiles)
    print u"[Info] 检测到处理文件:",str(total) ,u"处理成功:",str(total-failednum),u"处理失败：",str(failednum)
    print u"[Info] 导出路径为：",args.output
    if failednum!=0:
        print u"[Info] 处理失败文件清单保存至文件: ./failed.txt"
    print "[Info] Program completed!"


if __name__ == "__main__":
    main()
    pass